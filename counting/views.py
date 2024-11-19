from django.core.serializers import serialize
import decimal
import json
import random
from datetime import datetime
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import send_mail, BadHeaderError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.db.models import F
from django.utils.safestring import mark_safe
from django.core.mail import send_mail
from asor_accounting.settings import EMAIL_HOST_USER

from counting.functions import recalculatetotal, totalcalc, generate_unique_id
from .forms import *
from .models import *
from .models import *
from django.db.models import F, Sum
from openpyxl import Workbook
from xhtml2pdf import pisa
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from io import BytesIO
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from weasyprint import HTML


app_name = 'counting'


def login_request(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.info(request, f"You are now logged in as {username}.")
                return redirect("index")
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid username or password.")
    form = AuthenticationForm()
    return render(request, "../templates/registration/login.html", context={"login_form": form})


def register_request(request):
    if request.method == "POST":
        form = NewUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Registration successful.")
            return redirect("counting:index")
        messages.error(request, "Unsuccessful registration. Invalid information.")
    form = NewUserForm()
    return render(request=request, template_name="../templates/registration/register.html",
                  context={"register_form": form})


def index(request):
    return render(request, "counting/index.html")


def counting(request):
    if request.user.is_authenticated:
        count_id = request.session.get('count_id')
        unique_id = generate_unique_id(10)
        user = request.user.id
        all_donation_categories = donation_category.objects.all()
        all_parishioners = parishoners.objects.all().order_by('last_name')
        
        autosave_data = None
        
        if count_id is not None:
            count_exists = True
            try:
                autosave_data = autosave.objects.get(count_id=count_id)
            except autosave.DoesNotExist:
                autosave_data = None
        else:
            count_exists = False
            new_count = count()
            new_count.count_date = datetime.now()
            new_count.user_id = user
            new_count.count_type_id = 1
            new_count.save()
            count_id = new_count.pk
            request.session['count_id'] = new_count.pk
        
        context = {
            "all_donation_categories": all_donation_categories,
            "new_count_id": count_id,
            "all_parishioners": all_parishioners,
            "unique_id": unique_id,
            "count_exists": count_exists,
            "autosave_data": autosave_data.data if autosave_data else None  # Include autosave data if it exists
        }
        return render(request, "counting/count.html", context)
    else:
        messages.error(request, "You are not logged in.")
        return redirect('login')


def updatecount(request):
    count_id = request.POST.get('count_id')
    denomination = request.POST.get('denomination')
    cat_id = request.POST.get('category_id')
    custom_cat_id = request.POST.get('custom_category_id')
    new_value = int(request.POST.get('new_value'))
    if custom_cat_id is not None:
        if unnamed_donations.objects.filter(count_id=count_id, donation_date=datetime.now(), category_id=cat_id,
                                            custom_category=custom_cat_id).exists():
            exists = 1
            donation = unnamed_donations.objects.get(
                count_id=count_id,
                donation_date=datetime.now(),
                category_id=cat_id,
                custom_category = custom_cat_id
            )
        else:
            exists = 0
            donation = unnamed_donations()
        multiplier = 0

        match denomination:
            case '1':
                multiplier = 1
                donation.one_count = new_value
            case '2':
                multiplier = 2
                donation.two_count = new_value
            case '5':
                multiplier = 5
                donation.five_count = new_value
            case '10':
                multiplier = 10
                donation.ten_count = new_value
            case '20':
                multiplier = 20
                donation.twenty_count = new_value
            case '50':
                multiplier = 50
                donation.fifty_count = new_value
            case '100':
                multiplier = 100
                donation.hundred_count = new_value
        if exists == 1:
            donation.save()
            current_total = recalculatetotal(
                donation.one_count,
                donation.two_count,
                donation.five_count,
                donation.ten_count,
                donation.twenty_count,
                donation.fifty_count,
                donation.hundred_count
            )
            donation.donation_total = current_total
            donation.save()
        else:
            donation.donation_date = datetime.now()
            donation.count_id = count_id
            donation.category_id = cat_id
            if custom_cat_id is not None and custom_cat_id != '':
                custom_category_instance = donation_category_custom.objects.get(id=custom_cat_id)
                donation.custom_category = custom_category_instance
            donation.payment_type = payment_type.objects.get(id=1)
            current_total = multiplier * new_value
            donation.donation_total = current_total
            donation.save()

    else:
        if unnamed_donations.objects.filter(count_id=count_id, donation_date=datetime.now(), category_id=cat_id).exists():
            exists = 1
            donation = unnamed_donations.objects.get(
                count_id=count_id,
                donation_date=datetime.now(),
                category_id=cat_id
            )
        else:
            exists = 0
            donation = unnamed_donations()
        multiplier = 0
        match denomination:
            case '1':
                multiplier = 1
                donation.one_count = new_value
            case '2':
                multiplier = 2
                donation.two_count = new_value
            case '5':
                multiplier = 5
                donation.five_count = new_value
            case '10':
                multiplier = 10
                donation.ten_count = new_value
            case '20':
                multiplier = 20
                donation.twenty_count = new_value
            case '50':
                multiplier = 50
                donation.fifty_count = new_value
            case '100':
                multiplier = 100
                donation.hundred_count = new_value
        if exists == 1:
            donation.save()
            current_total = recalculatetotal(
                donation.one_count,
                donation.two_count,
                donation.five_count,
                donation.ten_count,
                donation.twenty_count,
                donation.fifty_count,
                donation.hundred_count
            )
            donation.donation_total = current_total
            donation.save()
        else:
            donation.donation_date = datetime.now()
            donation.count_id = count_id
            donation.category_id = cat_id
            if custom_cat_id is not None and custom_cat_id != '':
                custom_category_instance = donation_category_custom.objects.get(id=custom_cat_id)
                donation.custom_category = custom_category_instance
            donation.payment_type = payment_type.objects.get(id=1)
            current_total = multiplier * new_value
            donation.donation_total = current_total
            donation.save()
    count_ones = 0
    count_twos = 0
    count_fives = 0
    count_tens = 0
    count_twentys = 0
    count_fiftys = 0
    count_hundreds = 0
    for donations in unnamed_donations.objects.filter(category_id__in=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
                                                      count_id=count_id, donation_date=datetime.now()):
        count_ones += donations.one_count
        total_ones = totalcalc(count_ones, 1)
        count_twos += donations.two_count
        total_twos = totalcalc(count_twos, 2)
        count_fives += donations.five_count
        total_fives = totalcalc(count_fives, 5)
        count_tens += donations.ten_count
        total_tens = totalcalc(count_tens, 10)
        count_twentys += donations.twenty_count
        total_twentys = totalcalc(count_twentys, 20)
        count_fiftys += donations.fifty_count
        total_fiftys = totalcalc(count_fiftys, 50)
        count_hundreds += donations.hundred_count
        total_hundreds = totalcalc(count_hundreds, 100)
        grand_total = total_ones + total_twos + total_fives + total_tens + total_twentys + total_fiftys + total_hundreds
    response = {'current_total': current_total, 'current_count': new_value, 'denomination': denomination,
                'category_id': cat_id, 'count_ones': count_ones, 'count_twos': count_twos, 'count_fives': count_fives,
                'count_tens': count_tens, 'count_twentys': count_twentys, 'count_fiftys': count_fiftys,
                'count_hundreds': count_hundreds, 'total_ones': total_ones, 'total_twos': total_twos,
                'total_fives': total_fives, 'total_tens': total_tens, 'total_twentys': total_twentys,
                'total_fiftys': total_fiftys, 'total_hundreds': total_hundreds, 'grand_total': grand_total}
    return JsonResponse(response)


def getcheckdata(request):
    all_parishioners = parishoners.objects.all().order_by('last_name')
    return_string = ''
    return_string += '<td><input class="parishoner-check-number" type="text"></td><th><input ' \
                     'class="cash-donation" type="checkbox"></th> <td><select '
    return_string += 'class="parishioner-select" name="parishioner-choice"><option>Select Parishoner...</option>'
    return_string += '<option data-type="add">Add New Parishioner</option>'
    for person in all_parishioners:
        return_string += f'<option value="{person.id}">{person.last_name}, {person.first_name}</option>'
    return_string += '</select></td><td><input class="donation-amount" min="0" type="number"></td></tr>'
    response = {'return_string': return_string}
    return JsonResponse(response)


def getcontactdata(request):
    all_states = parishoner_state.objects.all()
    ctypes = contact_type.objects.all()
    state_text = ''
    contact_text = ''
    for state in all_states:
        state_text += f'<option value="{state.id}">{state.state_abbr}</option>'
    for contact in ctypes:
        if contact.id != 3:
            contact_text += f'<option value="{contact.id}">{contact.contact_name}</option>'
    response = {'state_text': state_text, 'contact_text': contact_text}
    return JsonResponse(response)


def savenewparishioner(request):
    response = {}
    first_name = request.POST.get('first_name')
    last_name = request.POST.get('last_name')
    full_name = last_name + ', ' + first_name
    new_parishioner = parishoners()
    new_parishioner.first_name = first_name
    new_parishioner.last_name = last_name
    new_parishioner.full_name = first_name + " " + last_name
    new_parishioner.save()
    save_address = request.POST.get('save_address')
    save_contact = request.POST.get('save_contact')
    response.update({'parishoner_id': new_parishioner.id})
    response.update({'parishoner_name': full_name})
    if save_address == 'true':
        address_1 = request.POST.get('address_1')
        address_2 = request.POST.get('address_2')
        city = request.POST.get('city')
        state = request.POST.get('state')
        zip = request.POST.get('zip')
        p_address = parishoner_address()
        p_address.parishoner = new_parishioner
        p_address.address_1 = address_1
        p_address.address_2 = address_2
        p_address.city = city
        p_address.state = parishoner_state.objects.get(pk=state)
        p_address.zip = zip
        p_address.save()
        response.update({'address': 'saved'})
    if save_contact == 'true':
        phone_type = request.POST.get('phone_type')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        p_contact = parishoner_contact()
        p_contact.parishoner = new_parishioner
        p_contact.contact_type = contact_type.objects.get(pk=phone_type)
        p_contact.contact_data = phone_number
        p_email = parishoner_contact()
        p_email.parishoner = new_parishioner
        p_email.contact_type = contact_type.objects.get(pk=3)
        p_email.contact_data = email
        p_contact.save()
        p_email.save()
        response.update({'contact': 'saved'})
    return JsonResponse(response)


# def donation_note(request):
#     category_ids = [5, 6, 9]
#     count_date = datetime.now()
#     special_categories_unamed = unnamed_donations.objects.filter(Q(category_id__in=category_ids) & Q(donation_date=count_date)).select_related('category')
#     special_categories_named = named_donations.objects.filter(Q(category_id__in=category_ids) & Q(donation_date=count_date)).select_related('category')
#     context = {'special_categories_unamed': special_categories_unamed, 'special_categories_named': special_categories_named}
#     if len(special_categories_unamed) > 0 and len(special_categories_named) > 0:
#         return render(request, "counting/review_count.html")
#     else:
#         return render(request, "counting/donation_note.html", context)
# # the query is empty

def submit_count(request):
    if request.method == 'POST':
        check_return = {}
        data = json.loads(request.body)
        return_object = data.get('return_object', {})
        count_id = return_object.get('count_id')  # Retrieve count_id from return_object
        named_donations.objects.filter(count_id=count_id).delete()
        checks = return_object.get('checks', [])  # Retrieve checks list from return_object
        for item_key, item in checks.items():
            new_named_donation = named_donations()
            new_named_donation.donation_date = datetime.now()
            new_named_donation.donation_total = item['donation_amount']
            category_id = item['category']
            category_instance = donation_category.objects.get(id=category_id)
            new_named_donation.category = category_instance
            custom_category_id = item['custom_category']
            if custom_category_id is not None:
                custom_category_instance = donation_category_custom.objects.get(id=custom_category_id)
                new_named_donation.custom_category = custom_category_instance
            count_instance = count.objects.get(id=count_id)  # Use the retrieved count_id
            new_named_donation.count = count_instance

            parishoner_id = item['parishioner']
            parishoner_instance = parishoners.objects.get(id=parishoner_id)
            new_named_donation.parishoner = parishoner_instance

            if item['cash_donation']:
                new_named_donation.payment_type_id = 1
                new_named_donation.donation_detail_id = None
            else:
                new_named_donation.payment_type_id = 2
                new_donation_detail = donation_detail()
                new_donation_detail.check_number = item['check_number']
                new_donation_detail.save()
                new_named_donation.donation_detail_id = new_donation_detail

            new_named_donation.save()
            check_id = new_named_donation.pk
            check_return[item_key] = check_id

        items_saved = {
            'Message': f'{len(checks)} checks saved successfully!',
            'saved_checks': check_return,
            'success': 1,
            'count_id': count_id,
        }
        return JsonResponse(items_saved)


def autosave_checks(request):
    if request.method == 'POST':
        check_return = {}
        data = json.loads(request.body)
        return_object = data.get('return_object', {})
        count_id = return_object.get('count_id')  # Retrieve count_id from return_object
        named_donations.objects.filter(count_id=count_id).delete()
        checks = return_object.get('checks', [])  # Retrieve checks list from return_object
        for item_key, item in checks.items():
            new_named_donation = named_donations()
            new_named_donation.donation_date = datetime.now()
            new_named_donation.donation_total = item['donation_amount']

            category_id = item['category']
            category_instance = donation_category.objects.get(id=category_id)
            new_named_donation.category = category_instance

            custom_category_id = item['custom_category']
            if custom_category_id is not None:
                custom_category_instance = donation_category_custom.objects.get(id=custom_category_id)
                new_named_donation.custom_category = custom_category_instance

            count_instance = count.objects.get(id=count_id)  # Use the retrieved count_id
            new_named_donation.count = count_instance

            parishoner_id = item['parishioner']
            parishoner_instance = parishoners.objects.get(id=parishoner_id)
            new_named_donation.parishoner = parishoner_instance

            if item['cash_donation']:
                new_named_donation.payment_type_id = 1
                new_named_donation.donation_detail_id = None
            else:
                new_named_donation.payment_type_id = 2
                new_donation_detail = donation_detail()
                new_donation_detail.check_number = item['check_number']
                new_donation_detail.save()
                new_named_donation.donation_detail_id = new_donation_detail

            new_named_donation.save()
            check_id = new_named_donation.pk
            check_return[item_key] = check_id

        items_saved = {
             'Message': f'{len(checks)} checks saved successfully!',
            'saved_checks': check_return,
            'success': 1,
            'count_id': count_id,
        }
        return JsonResponse(items_saved)


def review_count(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        count_id = data.get('count_id')
        count_instance = count.objects.get(id=count_id)

        named_donations_data = named_donations.objects.select_related('parishoner', 'category', 'payment_type').filter(
            count=count_instance)

        named_donations_data = named_donations_data.annotate(
            custom_category_name=F('custom_category__category_name'))

        unnamed_donations_data = unnamed_donations.objects.select_related('category', 'payment_type',
                                                                          'custom_category').filter(
            count=count_instance)
        unnamed_donations_data = unnamed_donations_data.annotate(
            custom_category_name=F('custom_category__category_name'))

        if request.user.is_authenticated:
            context = {
                "count_id": count_id,
                'count_instance': count_instance,
                'named_donations_data': named_donations_data,
                'unnamed_donations_data': unnamed_donations_data,
            }
            return render(request, "counting/review_count.html", context)
        else:
            messages.error(request, "You are not logged in.")


def create_custom_category(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        count_category = data.get('donation_category_type')
        custom_category_name = data.get('category_name')
        if data.get('unique_id') is not None and data.get('unique_id') != "":
            unique_id = data.get('unique_id')
        else:
            unique_id = ''
        category_instance = donation_category.objects.get(id=count_category)
        original_name = category_instance.category_name
        new_custom_category = donation_category_custom()
        new_custom_category.category_name = custom_category_name
        new_custom_category.category_type = category_instance
        new_custom_category.save()
        custom_category_id = new_custom_category.id
        response_data = {'custom_category_id': custom_category_id, 'count_category': count_category,
                         'category_name': custom_category_name, 'unique_id': unique_id, 'original_name': original_name,
                         'success': 1}
        return JsonResponse(response_data)


def count_report_redirect(request):
    if request.method == 'POST':
        # Your logic to determine the URL of the new page
        new_page_url = '/count_report'

        # Return the URL in a JSON response
        return JsonResponse({'new_page_url': new_page_url})
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


def count_report(request):
    if request.user.is_authenticated:
        first_name = request.user.first_name
        last_name = request.user.last_name
        users_name = f"{first_name} {last_name}"
        count_date = datetime.now()
        count_id = request.session.get('count_id')
        membership_donation_id = 10

        # category donation data with QB account
        category_donation_data = {}

        # Get the unnamed donations related to the specified count ID
        unnamed_donations_data = unnamed_donations.objects.filter(count_id=count_id).select_related('category')
        named_donations_data = named_donations.objects.filter(count_id=count_id).select_related('category')
        
        for donation in unnamed_donations_data:
            if donation.custom_category:
                category_name = donation.custom_category.category_name
                category_qb_account = donation.custom_category.category_type.category_qb_account if donation.custom_category.category_type.category_qb_account else "N/A"
            else:
                category_name = donation.category.category_name
                category_qb_account = donation.category.category_qb_account if donation.category.category_qb_account else "N/A"

            donation_key = (category_qb_account, category_name)
            category_donation_data[donation_key] = category_donation_data.get(donation_key, 0) + donation.donation_total

        # Repeat similar logic for named donations
        for donation in named_donations_data:
            if donation.custom_category:
                category_name = donation.custom_category.category_name
                category_qb_account = donation.custom_category.category_type.category_qb_account if donation.custom_category.category_type.category_qb_account else "N/A"
            else:
                category_name = donation.category.category_name
                category_qb_account = donation.category.category_qb_account if donation.category.category_qb_account else "N/A"

            donation_key = (category_qb_account, category_name)
            category_donation_data[donation_key] = category_donation_data.get(donation_key, 0) + donation.donation_total

        # Membership data
        membership_data = {}
        named_donations_data = named_donations.objects.filter(
            category__id=membership_donation_id,
            count__id=count_id
        ).values(
            'parishoner__full_name',
            'payment_type__payment_type_name',
            'donation_total'
        )

        for index, donation in enumerate(named_donations_data, start=1):
            parishoner_name = donation['parishoner__full_name']
            payment_type = donation['payment_type__payment_type_name']
            donation_amount = donation['donation_total']
            membership_data[index] = {
                'parishoner_name': parishoner_name,
                'payment_type': payment_type,
                'donation_amount': donation_amount
            }

        # Denomination data calculation
        denomination_data = {}
        denomination_aggregates = unnamed_donations.objects.filter(count__id=count_id).aggregate(
            one_bills_count=Sum('one_count'),
            two_bills_count=Sum('two_count'),
            five_bills_count=Sum('five_count'),
            ten_bills_count=Sum('ten_count'),
            twenty_bills_count=Sum('twenty_count'),
            fifty_bills_count=Sum('fifty_count'),
            hundred_bills_count=Sum('hundred_count'),
            one_bills_value=Sum(F('one_count') * 1),
            two_bills_value=Sum(F('two_count') * 2),
            five_bills_value=Sum(F('five_count') * 5),
            ten_bills_value=Sum(F('ten_count') * 10),
            twenty_bills_value=Sum(F('twenty_count') * 20),
            fifty_bills_value=Sum(F('fifty_count') * 50),
            hundred_bills_value=Sum(F('hundred_count') * 100),
        )

        # Define the denominations and their corresponding fields in the unnamed_donations_data
        denominations = {
            "$1 Bills": ["one_bills_count", "one_bills_value"],
            "$2 Bills": ["two_bills_count", "two_bills_value"],
            "$5 Bills": ["five_bills_count", "five_bills_value"],
            "$10 Bills": ["ten_bills_count", "ten_bills_value"],
            "$20 Bills": ["twenty_bills_count", "twenty_bills_value"],
            "$50 Bills": ["fifty_bills_count", "fifty_bills_value"],
            "$100 Bills": ["hundred_bills_count", "hundred_bills_value"],
        }

        for denomination, (count_field, value_field) in denominations.items():
            bills_count = denomination_aggregates.get(count_field, 0)
            bills_value = denomination_aggregates.get(value_field, 0)
            denomination_data[denomination] = {
                "bills_count": bills_count,
                "bills_value": bills_value,
            }

        # Grand total calculations
        check_grand_total = named_donations.objects.filter(count__id=count_id).aggregate(
            grand_total=Sum('donation_total')
        )['grand_total'] or 0

        cash_grand_total = unnamed_donations.objects.filter(count__id=count_id).aggregate(
            grand_total=Sum(F('one_count') + 2 * F('two_count') + 5 * F('five_count') +
                            10 * F('ten_count') + 20 * F('twenty_count') + 50 * F('fifty_count') +
                            100 * F('hundred_count'))
        )['grand_total'] or 0

        grand_total = check_grand_total + cash_grand_total
        cash_grand_total = "{:.2f}".format(cash_grand_total)
        check_grand_total = "{:.2f}".format(check_grand_total)
        grand_total = "{:.2f}".format(grand_total)

        context = {
            "category_donation_data": category_donation_data,
            "count_date": count_date,
            "users_name": users_name,
            "membership_data": membership_data,
            "denomination_data": denomination_data,
            'check_grand_total': check_grand_total,
            'cash_grand_total': cash_grand_total,
            'grand_total': grand_total
        }

        # Generate and send reports
        subject_date = count_date.strftime("%m/%d/%Y")
        subject = 'Donation Count for ' + subject_date
        message = 'Here is the count from today'
        recipient_list = ['christopherjoellakey@gmail.com', 'bhend60@gmail.com', 'atimofeeva@aol.com', 'mariyabobo303@outlook.com', 'asorrocadenver@gmail.com']
        excel_name = 'Count_Workbook_' + subject_date + '.xlsx'
        pdf_name = 'Count_PDF_' + subject_date + '.pdf'

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        yellow_fill = PatternFill(start_color='FFFF00',
                                  end_color='FFFF00',
                                  fill_type='solid')

        ws.append(['Count Report', '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row, column=1).fill = yellow_fill

        ws.append(['Count Date ' + subject_date, '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')

        ws.append(['Counted By: ' + users_name, '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')

        ws.append(['Category Donations Data', '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row, column=1).fill = yellow_fill

        ws.append(['QuickBooks Account', 'Category', 'Total Donation'])
        # Applying styles to headers
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.fill = yellow_fill
            cell.border = thin_border
        
        currency_format = '"$"#,##0.00'

        # Formatting and writing each row of category data
        for (category_name, category_qb_account), total_donation in category_donation_data.items():
            ws.append([category_name, category_qb_account, total_donation])
            last_row = ws.max_row
            ws.cell(row=last_row, column=3).number_format = '"$"#,##0.00'
            for cell in ws[last_row]:
                cell.border = thin_border

        ws.append(['Membership Data', '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row, column=1).fill = yellow_fill

        data2 = [
            ['Parishoner Name', 'Payment Type', 'Donation Amount']
        ]

        for row in data2:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = yellow_fill

        for index, data in membership_data.items():
            row = [
                data['parishoner_name'],
                data['payment_type'],
                data['donation_amount']
            ]
            ws.append(row)
            last_row = ws.max_row
            ws.cell(row=last_row, column=3).number_format = currency_format
            for cell in ws[last_row]:
                cell.border = thin_border

        ws.append(['Cash Counts', '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row, column=1).fill = yellow_fill

        data2 = [
            ['Denomination', 'Bill Count', 'Denomination Total']
        ]

        for row in data2:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = yellow_fill

        for denomination, data in denomination_data.items():
            ws.append([denomination, data['bills_count'], data['bills_value']])
            last_row = ws.max_row  # Get the last row number
            ws.cell(row=last_row, column=3).number_format = currency_format

        ws.append(['Donation Totals', '', ''])
        last_row = ws.max_row
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row, column=1).fill = yellow_fill

        data3 = [
            ['Checks Total', check_grand_total],
            ['Cash Total', cash_grand_total],
            ['Grand Total', grand_total]
        ]

        for row_data in data3:
            ws.append(row_data)
            last_row = ws.max_row
            ws.cell(row=last_row, column=1).fill = yellow_fill
            ws.cell(row=last_row, column=2).number_format = currency_format

        max_length = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    max_length[cell.column] = max(
                        (len(str(cell.value)) + 2) * 1.2,
                        max_length.get(cell.column, 0)
                    )

        for column, length in max_length.items():
            ws.column_dimensions[get_column_letter(column)].width = length

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        html_content = render_to_string('counting/count_report.html', context)
        pdf_output = BytesIO()
        HTML(string=html_content).write_pdf(pdf_output)
        pdf_output.seek(0)
        pdf_content = pdf_output.getvalue()

        wb.save(output)
        excel_content = output.getvalue()
        email = EmailMessage(
            subject,
            message,
            EMAIL_HOST_USER,
            recipient_list,
        )
        email.attach(excel_name, excel_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.attach(pdf_name, pdf_content, 'application/pdf')
        email.send()

        if 'count_id' in request.session:
            del request.session['count_id']
            request.session.save()

        return render(request, "counting/count_report.html", context)
    else:
        messages.error(request, "You are not logged in.")

def remove_count_session(request):
    if 'count_id' in request.session:
        del request.session['count_id']
        request.session.save()
        context = {
            "success": 1
        }
        return JsonResponse(context)

def autosave_count(request):
    if request.method == 'POST':
        post_data = json.loads(request.body)
        count_id = post_data.get('count_id')
        count_data = post_data.get('data')

        # Ensure the count exists
        try:
            count_instance = count.objects.get(pk=count_id)
        except count.DoesNotExist:
            return JsonResponse({'error': 'Count does not exist'}, status=404)

        # Update or create autosave instance for the count
        autosave_instance, created = autosave.objects.update_or_create(
            count=count_instance,
            defaults={'data': count_data},
        )

        response_message = 'Data saved successfully!' if created else 'Data updated successfully!'
        return JsonResponse({'message': response_message, 'success': 1, 'count_id': count_id})

